"""Read-only source adapters. They emit source records; only the ingestion service writes KB rows."""
from __future__ import annotations

import asyncio
import csv
from dataclasses import dataclass
from datetime import UTC, datetime
from io import StringIO
from pathlib import Path
from typing import ClassVar

from docx import Document
from openpyxl import load_workbook

from carrier_kb.ingest.registry import SourceDefinition


@dataclass(frozen=True)
class CapturedRecord:
    native_id: str
    body: str
    source_url: str | None
    occurred_at: datetime
    metadata: dict


class SourceAdapter:
    async def capture(self, source: SourceDefinition) -> list[CapturedRecord]:
        raise NotImplementedError


class GoogleDriveAdapter(SourceAdapter):
    """Fetches only registry file IDs via Docs/Drive APIs; folder-wide discovery is intentionally absent."""
    async def capture(self, source: SourceDefinition) -> list[CapturedRecord]:
        # Wire Google credentials/client here after granting the service account access to exact files.
        raise NotImplementedError("Google Drive adapter requires approved service-account configuration")


class SlackAdapter(SourceAdapter):
    """Fetches only registry channel IDs. Slack search is prohibited because it bypasses the allowlist."""
    async def capture(self, source: SourceDefinition) -> list[CapturedRecord]:
        raise NotImplementedError("Slack adapter requires approved bot-token configuration")


class LohiViewAdapter(SourceAdapter):
    """Reads migration-owned views/dossiers only; no natural-language or model-supplied SQL exists here."""
    async def capture(self, source: SourceDefinition) -> list[CapturedRecord]:
        if not source.view:
            raise ValueError("LoHi source is missing its approved view")
        raise NotImplementedError("LoHi view adapter requires the read-only dossier connection")


class StaticFileAdapter(SourceAdapter):
    """Explicitly listed UTF-8 files for interim, reviewable corpus seeding."""

    async def capture(self, source: SourceDefinition) -> list[CapturedRecord]:
        if not source.path:
            raise ValueError("static source is missing a path")
        path = Path(source.path)
        body = self._read_body(path)
        stat = path.stat()
        chunks = self._chunks(body, source.chunk_chars)
        return [CapturedRecord(
            native_id=f"{path.resolve()}#chunk-{index}", body=chunk, source_url=None,
            occurred_at=datetime.fromtimestamp(stat.st_mtime, tz=UTC),
            metadata={"filename": path.name, "chunk_index": index, "chunk_count": len(chunks)},
        ) for index, chunk in enumerate(chunks)]

    @staticmethod
    def _read_body(path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in {".txt", ".md"}:
            return path.read_text(encoding="utf-8")
        if suffix == ".csv":
            rows = csv.reader(StringIO(path.read_text(encoding="utf-8")))
            return "\n".join(" | ".join(cell.strip() for cell in row) for row in rows)
        if suffix == ".docx":
            return "\n".join(p.text.strip() for p in Document(path).paragraphs if p.text.strip())
        if suffix == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            return "\n".join(
                " | ".join(str(cell) if cell is not None else "" for cell in row)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows(values_only=True)
            )
        raise ValueError(f"unsupported static file type: {suffix}")

    @staticmethod
    def _chunks(body: str, chunk_chars: int) -> list[str]:
        lines = body.splitlines()
        chunks: list[str] = []
        current: list[str] = []
        size = 0
        for line in lines:
            if current and size + len(line) + 1 > chunk_chars:
                chunks.append("\n".join(current))
                current, size = [], 0
            current.append(line)
            size += len(line) + 1
        if current:
            chunks.append("\n".join(current))
        return chunks or [""]


class FrontCsvAdapter(SourceAdapter):
    """Curate Front messages into conversation-level, PII-minimized records."""

    TERMS: ClassVar[tuple[str, ...]] = (
        "onboard", "application", "packet", "rmis", "insurance", "coi", "tenstreet",
        "training", "driver", "factoring", "payment", "toll", "trailer", "power only",
        "lane", "load", "market", "status", "document", "w9", "h2s", "safeland",
    )
    EXCLUDE_TERMS: ClassVar[tuple[str, ...]] = ("meta ads receipt", "craigslist post", "paid posting")
    TOPICS: ClassVar[dict[str, tuple[str, ...]]] = {
        "verification": ("rmis", "evident", "insurance", "coi"),
        "training": ("training", "tenstreet", "safeland", "w9", "h2s"),
        "payment": ("payment", "paid", "factoring", "rate", "invoice"),
        "equipment": ("equipment", "trailer", "power only", "flatbed"),
        "status": ("status", "application", "onboard", "packet", "document"),
        "escalation": ("contact", "follow up", "follow-up", "stuck", "help"),
    }

    async def capture(self, source: SourceDefinition) -> list[CapturedRecord]:
        if not source.path:
            raise ValueError("Front source is missing a path")
        path = Path(source.path)
        groups: dict[str, list[str]] = {}
        for row in await asyncio.to_thread(self._read_rows, path):
            text = " ".join((row.get(key) or "") for key in ("Subject", "Extract", "Tags"))
            lowered = text.lower()
            if (
                not any(term in lowered for term in self.TERMS)
                or any(term in lowered for term in self.EXCLUDE_TERMS)
            ):
                continue
            conversation_id = row.get("Conversation ID") or row.get("Message ID") or "unknown"
            topics = [topic for topic, terms in self.TOPICS.items() if any(term in lowered for term in terms)]
            groups.setdefault(conversation_id, []).append(f"topics: {', '.join(topics)}\n{text.strip()}")
        stat = path.stat()
        occurred_at = datetime.fromtimestamp(stat.st_mtime, tz=UTC)
        records: list[CapturedRecord] = []
        for conversation_id, messages in groups.items():
            body = "\n".join(dict.fromkeys(message for message in messages if message))
            records.append(CapturedRecord(
                native_id=conversation_id, body=body, source_url=None, occurred_at=occurred_at,
                metadata={"conversation_id": conversation_id, "message_count": len(messages)},
            ))
        return records

    @staticmethod
    def _read_rows(path: Path) -> list[dict[str, str]]:
        with path.open(newline="", encoding="utf-8") as handle:
            return list(csv.DictReader(handle))
